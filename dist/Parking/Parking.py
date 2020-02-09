from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.button import Button
from kivy.properties import ObjectProperty, StringProperty
from kivy.factory import Factory
from kivy.lang.builder import Builder
from kivy.clock import Clock
from functools import partial
from datetime import date, datetime
from shutil import copy
import locale
import xlrd, unicodedata
import openpyxl
import os, sys, traceback
import re


import configparser
config = configparser.ConfigParser()
config.read('config.ini')

if not config['Default']['width'] or not config['Default']['height']:
    import ctypes
    user32 = ctypes.windll.user32
    screensize = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
    config['Default']['width'] = str(user32.GetSystemMetrics(0)-200)
    config['Default']['height'] = str(user32.GetSystemMetrics(1)-150)
    with open('config.ini', 'w') as configfile: #save
        config.write(configfile)

import sys
sys.setrecursionlimit(5000)

from kivy.core.window import Window
Window.clearcolor = (0.5, 0.5, 0.5, 1)

from kivy.config import Config
Config.set('graphics', 'width', config['Default']['width'])
Config.set('graphics', 'height', config['Default']['height'])
Config.set('graphics','borderless', 0)
Config.set('graphics','resizable', 1)
#Config.set('input', 'mouse', 'mouse,multitouch_on_demand')
Config.set('kivy', 'exit_on_escape', '1')
Config.write()

MyApp=None
class UI(BoxLayout):

    def __init__(self, **kwargs):
        super(UI, self).__init__(**kwargs)

    buscar_user = ObjectProperty()
    user_results = ObjectProperty()
    user_results_empty = ObjectProperty()
    label_nenhum_resultado = ObjectProperty()
    nome_doc_placas1 = ObjectProperty()
    nome_doc_placas2 = ObjectProperty()
    nome_doc_placas3 = ObjectProperty()

    buscar_aloc = ObjectProperty()
    aloc_results = ObjectProperty()
    aloc_results_empty = ObjectProperty()
    label_nenhum_resultado_aloc = ObjectProperty()
    aloc_label1 = ObjectProperty()
    aloc_label2 = ObjectProperty()
    aloc_label3 = ObjectProperty()

    user_sheet_path = 'planilhas/usuarios.xlsx'
    all_sheets_user = None
    user_sheet = None
    aloc_sheet_path = 'planilhas/alocamentos.xlsx'
    all_sheets_aloc = None
    aloc_sheet = None

    ## FUNCOES COMUNS ##
    def _refocus_text_input(self, field, instance):
        field.focus=True

    def set_cursor(self, cursor, instance):
        cursor.do_cursor_movement('cursor_home')

    def remove_accents(self, input_str):
        nfkd_form = unicodedata.normalize('NFKD', input_str)
        only_ascii = nfkd_form.encode('ASCII', 'ignore')
        return only_ascii

    def sheet_len(self, ws):
        res = 0
        for row in ws.iter_rows():
            x = [cell.value for cell in row]
            if not all(s == None for s in x):
                res += 1
        return res

    def iter_rows(self, ws):
        row_values = []
        for row in ws.iter_rows():
            x = [cell.value for cell in row]
            if not all(s == None for s in x):
                row_values.append(x)
        return reversed(row_values)

    def convert_to_string(self, vlist):
        new_list = []
        for v in vlist:
            if isinstance(v, float):
                new_list.append(str(int(v)))
            elif not v:
                new_list.append('')
            else:
                new_list.append(str(v))
        return new_list

    def deletar(self, result, sheet, instance):
        box = BoxLayout(orientation='horizontal', spacing=5)
        dele = Button(text='DELETAR', height='50sp', size_hint=(0.5, None), background_color=(1,0,0,1))#, foreground_color=(1,0,0,1))
        canc = Button(text='Cancelar', height='50sp', size_hint=(0.5, None))
        box.add_widget(dele)
        box.add_widget(canc)
        popup_deletar = Popup(title='Deletar',
                              content=box,
                              size_hint=(0.3, 0.2))
        canc.on_release = popup_deletar.dismiss
        dele.on_release = partial(self.del_pop_confirmed, result, popup_deletar, sheet)
        popup_deletar.open()

    def del_pop_confirmed(self, result, popup_deletar, sheet):
        if sheet == 'user':
            self.carregar_sheet_user()
            self.user_sheet.delete_rows(int(result[4]))
            self.all_sheets_user.save(self.user_sheet_path)
            self.load_usertable_results(self.buscar_user.text)
            self.fechar_sheet_user()
        elif sheet == 'aloc':
            self.carregar_sheet_aloc()
            self.aloc_sheet.delete_rows(int(result[7]))
            self.all_sheets_aloc.save(self.aloc_sheet_path)
            self.load_aloctable_results(self.buscar_aloc.text)
            self.fechar_sheet_aloc()

        popup_deletar.dismiss()
    ########################################

    ### MANIPULACAO PLANILHA DE USUARIOS ###
    def carregar_sheet_user(self):
        self.all_sheets_user = openpyxl.load_workbook(self.user_sheet_path)
        self.user_sheet = self.all_sheets_user[self.all_sheets_user.sheetnames[0]]

    def fechar_sheet_user(self):
        self.all_sheets_user.close()

    def clear_usertable_results(self):
        if self.user_results:
            self.user_results.clear_widgets()
        if self.user_results_empty:
            self.user_results_empty.clear_widgets()

        self.label_nenhum_resultado.text= ''
        self.nome_doc_placas1.text= ''
        self.nome_doc_placas2.text= ''
        self.nome_doc_placas3.text= ''

    def load_usertable_results(self, string_to_search):
        if string_to_search == 'limpar':
            self.clear_usertable_results()
            Clock.schedule_once(partial(self._refocus_text_input, self.buscar_user), 0)
            return

        results = self.procurar_user(string_to_search)
        self.clear_usertable_results()

        if results:
            self.nome_doc_placas1.text= 'Nome'
            self.nome_doc_placas2.text= 'Doc'
            self.nome_doc_placas3.text= 'Placas'

            for result in results:
                user = BoxLayout(orientation= 'horizontal',
                                 height= '30sp',
                                 size_hint= (1, None),
                                 spacing= 1)

                if result[0] == 'Limite de resultados excedido. (50)':
                    limite = Label(text= result[0],
                                   height= '30sp',
                                   size_hint= (1, None))
                    user.add_widget(limite)

                elif result[0] == 'Algo errado aconteceu. Tente novamente.':
                    erro = Label(text= result[0],
                                 height= '30sp',
                                 size_hint= (1, None))
                    user.add_widget(erro)
                    self.nome_doc_placas1.text= ''
                    self.nome_doc_placas2.text= ''
                    self.nome_doc_placas3.text= ''

                else:
                    deletar = Button(text= 'X',
                                     size_hint= (0.05, None),
                                     font_size= '14sp',
                                     height= '30sp',
                                     on_release= partial(self.deletar, result, 'user'))
                    nome = TextInput(text= result[0],
                                     write_tab= False,
                                     multiline= False,
                                     height= '30sp',
                                     size_hint= (0.3, None),
                                     background_normal= '',
                                     background_active= '',
                                     cursor_blink= False,
                                     cursor_color= (0,0,0,0),
                                     readonly= True,
                                     on_double_tap= partial(self.show_user, result))
                    doc = TextInput(text= result[1],
                                    write_tab= False,
                                    multiline= False,
                                    height= '30sp',
                                    size_hint= (0.3, None),
                                    background_normal= '',
                                    background_active= '',
                                    cursor_blink= False,
                                    cursor_color= (0,0,0,0),
                                    readonly= True,
                                    on_double_tap= partial(self.show_user, result))
                    placas = TextInput(text= result[2],
                                       write_tab= False,
                                       multiline= False,
                                       height= '30sp',
                                       size_hint= (0.3, None),
                                       background_normal= '',
                                       background_active= '',
                                       cursor_blink= False,
                                       cursor_color= (0,0,0,0),
                                       readonly= True,
                                       on_double_tap= partial(self.show_user, result))
                    alocar = Button(text= 'Alocar',
                                    size_hint= (0.095, None),
                                    font_size= '14sp',
                                    height= '30sp',
                                    #background_color= (0.4, 0.4, 0.4, 1),
                                    #background_normal= '',
                                    on_release= partial(self.criar_aloc, result))
                    Clock.schedule_once(partial(self.set_cursor, nome), 0)
                    Clock.schedule_once(partial(self.set_cursor, doc), 0)
                    Clock.schedule_once(partial(self.set_cursor, placas), 0)
                    user.add_widget(deletar)
                    user.add_widget(nome)
                    user.add_widget(doc)
                    user.add_widget(placas)
                    user.add_widget(alocar)
                self.user_results.add_widget(user)
                
        else:
            self.label_nenhum_resultado.text= 'Nenhum resultado encontrado!'
        Clock.schedule_once(partial(self._refocus_text_input, self.buscar_user), 0)

    def procurar_user(self, string):
        results = []
        try:
            self.carregar_sheet_user()
            for row, values in enumerate((self.iter_rows(self.user_sheet))):
                row_values = self.convert_to_string(values)
                row_values.append(str(self.sheet_len(self.user_sheet)-row))

                if len(results) > 50:
                    results.append(['Limite de resultados excedido. (50)'])
                    break
                if string == '' and row_values:
                    results.append(row_values)
                else:
                    for col in row_values:
                        if self.remove_accents(string.lower()) in self.remove_accents(col.lower()):
                            results.append(row_values)
                            break
            self.fechar_sheet_user()
        except Exception as e:
            traceback.print_exc(file=sys.stdout)
            msg = Label(text='Se estiver com a planilha aberta, feche e tente novamente.')
            popup_error = Popup(title='Erro',
                                content=msg,
                                size_hint=(0.5, 0.25))
            popup_error.open()

        return results

    def show_user(self, result, instance):
        content = UserDialog(user_nome=result[0], user_doc=result[1], user_placas='123', data_cadastro='Data de cadastro: '+result[3], linha_planilha='linha da planilha: '+result[4])

        content.ids.user_placas.text = result[2] #bug
        popup = Popup(title='Info', content=content, 
                      size=(600,300), size_hint=(None, None))

        content.alocar = partial(self.criar_aloc, result, content)
        content.salvar_user = partial(content.salvar_user, result, popup, content)
        content.cancel = popup.dismiss

        popup.open()

    def adicionar_usuario(self):
        content = NovoUser(data_cadastro=str('Data de cadastro: '+datetime.now().strftime("%d/%m/%Y %H:%M:%S")))
        content.user_nome.focus = True
        content.user_nome.text = ''
        content.user_doc.text = ''
        content.user_placas.text = ''

        popup = Popup(title='Novo Cliente', content=content, 
                      size=(600,275), size_hint=(None, None))

        content.add_user.on_release=partial(content.adicionar, popup)
        content.cancel.on_release = popup.dismiss
        popup.open()

    ###########################################
    ### MANIPULACAO PLANILHA DE ALOCAMENTOS ###
    ###########################################
    def carregar_sheet_aloc(self):
        self.all_sheets_aloc = openpyxl.load_workbook(self.aloc_sheet_path)
        self.aloc_sheet = self.all_sheets_aloc[self.all_sheets_aloc.sheetnames[0]]

    def fechar_sheet_aloc(self):
        self.all_sheets_aloc.close()

    def clear_aloctable_results(self):
        if self.aloc_results:
            self.aloc_results.clear_widgets()
        if self.aloc_results_empty:
            self.aloc_results_empty.clear_widgets()

        self.label_nenhum_resultado_aloc.text= ''
        self.aloc_label1.text= ''
        self.aloc_label2.text= ''
        self.aloc_label3.text= ''

    def load_aloctable_results(self, string_to_search):
        if string_to_search == 'limpar':
            self.clear_aloctable_results()
            Clock.schedule_once(partial(self._refocus_text_input, self.buscar_aloc), 0)
            return

        results = self.procurar_aloc(string_to_search)
        self.clear_aloctable_results()

        if results:
            self.aloc_label1.text= 'Nome'
            self.aloc_label2.text= 'Doc'
            self.aloc_label3.text= 'Placas'

            for result in results:
                user = BoxLayout(orientation= 'horizontal',
                                 height= '30sp',
                                 size_hint= (1, None),
                                 spacing= 1)

                if result[0] == 'Limite de resultados excedido. (50)':
                    limite = Label(text= result[0],
                                   height= '30sp',
                                   size_hint= (1, None))
                    user.add_widget(limite)

                elif result[0] == 'Algo errado aconteceu. Tente novamente.':
                    erro = Label(text= result[0],
                                 height= '30sp',
                                 size_hint= (1, None))
                    user.add_widget(erro)
                    self.aloc_label1.text= ''
                    self.aloc_label2.text= ''
                    self.aloc_label3.text= ''

                else:
                    deletar = Button(text= 'X',
                                     size_hint= (0.05, None),
                                     font_size= '14sp',
                                     height= '30sp',
                                     color= (1,1,1,1),
                                     background_color= (0.3,0,0,1),
                                     background_normal= '',
                                     on_release= partial(self.deletar, result, 'aloc'))
                    nome = TextInput(text= result[0],
                                     write_tab= False,
                                     multiline= False,
                                     height= '30sp',
                                     size_hint= (0.3, None),
                                     readonly= True,
                                     foreground_color= (1,1,1,1),
                                     background_color= (0.6,0,0,1) if result[6]=='NÃO' else (0,0.4,0,1),
                                     background_normal= '',
                                     background_active= '',
                                     cursor_blink= False,
                                     cursor_color= (0,0,0,0),
                                     on_double_tap= partial(self.show_aloc, result))
                    doc = TextInput(text= result[1],
                                    write_tab= False,
                                    multiline= False,
                                    height= '30sp',
                                    size_hint= (0.3, None),
                                    readonly= True,
                                    foreground_color= (1,1,1,1),
                                    background_color= (0.6,0,0,1) if result[6]=='NÃO' else (0,0.4,0,1),
                                    background_normal= '',
                                    background_active= '',
                                    cursor_blink= False,
                                    cursor_color= (0,0,0,0),
                                    on_double_tap= partial(self.show_aloc, result))
                    placas = TextInput(text= result[2],
                                       write_tab= False,
                                       multiline= False,
                                       height= '30sp',
                                       size_hint= (0.3, None),
                                       readonly= True,
                                       foreground_color= (1,1,1,1),
                                       background_color= (0.6,0,0,1) if result[6]=='NÃO' else (0,0.4,0,1),
                                       background_normal= '',
                                       background_active= '',
                                       cursor_blink= False,
                                       cursor_color= (0,0,0,0),
                                       on_double_tap= partial(self.show_aloc, result))
                    liberar = Button(text= 'Liberar',
                                     size_hint= (0.095, None),
                                     font_size= '14sp',
                                     height= '30sp',
                                     color= (0,0,0,1) if result[6]=='SIM' else (1,1,1,1),
                                     background_color= (0,0,0,1) if result[6]=='SIM' else (0,0.4,0,1),
                                     background_normal= '',
                                     disabled= True if result[6]=='SIM' else False,
                                     on_release= partial(self.show_aloc, result))
                    Clock.schedule_once(partial(self.set_cursor, nome), 0)
                    Clock.schedule_once(partial(self.set_cursor, doc), 0)
                    Clock.schedule_once(partial(self.set_cursor, placas), 0)
                    user.add_widget(deletar)
                    user.add_widget(nome)
                    user.add_widget(doc)
                    user.add_widget(placas)
                    user.add_widget(liberar)
                self.aloc_results.add_widget(user)
                
        else:
            self.label_nenhum_resultado_aloc.text= 'Nenhum resultado encontrado!'
        Clock.schedule_once(partial(self._refocus_text_input, self.buscar_aloc), 0)

    def procurar_aloc(self, string):
        results = []
        try:
            self.carregar_sheet_aloc()

            for row, values in enumerate((self.iter_rows(self.aloc_sheet))):
                row_values = self.convert_to_string(values)
                row_values.append(str(self.sheet_len(self.aloc_sheet)-row))

                if len(results) > 50:
                    results.append(['Limite de resultados excedido. (50)'])
                    break
                if string == '' and row_values:
                    results.append(row_values)
                else:
                    for col in row_values:
                        if self.remove_accents(string.lower()) in self.remove_accents(col.lower()):
                            results.append(row_values)
                            break
            self.fechar_sheet_aloc()
        except Exception as e:
            traceback.print_exc(file=sys.stdout)
            msg = Label(text='Se estiver com a planilha aberta, feche e tente novamente.')
            popup_error = Popup(title='Erro',
                                content=msg,
                                size_hint=(0.5, 0.25))
            popup_error.open()

        results = sorted(results, key = lambda x: x[6])
        return results

    def show_aloc(self, result, instance):
        data_saida=str(7*' ')+'Saída:\nNão Liberado!'
        if result[4]:
            data_saida=str(14*' ')+'Saída:\n'+result[4]

        if result[6] == 'SIM':
            content = AlocDialog(user_nome=result[0], user_doc=result[1], user_placas=result[2], data_entrada=str(12*' ')+'Entrada:\n'+result[3], data_saida=data_saida, liberado=result[6], linha_planilha='linha da planilha: '+result[7])
            content.preco.text = result[5]
            content.ids.preco.disabled = True
            content.ids.liberado.disabled = True
            content.salvar_aloc.disabled = True
            content.digitos = result[5][1:]
            if content.digitos[-2:] == '00':
                content.digitos = content.digitos[:-3]
            elif content.digitos[-1:] == '0':
                content.digitos = content.digitos[:-1]
        else:
            content = AlocDialog(user_nome=result[0], user_doc=result[1], user_placas=result[2], data_entrada=str(12*' ')+'Entrada:\n'+result[3], data_saida=data_saida, liberado=result[6], linha_planilha='linha da planilha: '+result[7])
            content.preco.focus = True
            content.preco.text = result[5]
            content.digitos = result[5][1:]
            if content.digitos[-2:] == '00':
                content.digitos = content.digitos[:-3]
            elif content.digitos[-1:] == '0':
                content.digitos = content.digitos[:-1]

        popup = Popup(title='Info', content=content,
                      size=(600,290), size_hint=(None, None))

        content.salvar_aloc.on_release = partial(content.liberar_aloc, popup)
        content.cancel.on_release = popup.dismiss

        popup.open()

    def criar_aloc(self, result, instance):
        lista = ['Nenhum']
        if all(i for i in result[2].split(';')):
            lista.extend(result[2].split(';'))

        content = NovoAloc(user_nome=result[0], user_doc=result[1], user_placas=lista, data_entrada=str(12*' ')+'Entrada:\n'+datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        content.ids.user_placas.text = 'Escolha!'
        content.preco.focus = True
        content.preco.text = '$0.00'

        popup = Popup(title='Info', content=content,
                      size=(500,280), size_hint=(None, None))

        content.criar_aloc.on_release = partial(content.adicionar_aloc, popup)
        popup.open()
        event = Clock.schedule_interval(partial(self.atualiza_entrada, content), 1)
        content.cancel.on_release = partial(self.unclock, event, popup)

    def atualiza_entrada(self, content, instance):
        content.data_entrada=str(12*' ')+'Entrada:\n'+datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    def unclock(self, event, popup):
        Clock.unschedule(event)
        popup.dismiss()

    def backup(self):
        content = BoxLayout(orientation='vertical', spacing=20)

        status = Label(text='Escolha o caminho e aperte OK para copiar arquivos.', size_hint=(1, None), height='20sp')
        path = TextInput(text=os.path.expanduser("~\Desktop"), size_hint=(1, None), height='30sp')
        botao_ok = Button(text='OK', size_hint=(0.2, None), height='30sp', pos_hint= {'right': 0.6})
        botao_ok.on_release=partial(self.backup_copy, path, status)

        content.add_widget(status)
        content.add_widget(path)
        content.add_widget(botao_ok)

        popup = Popup(title='Backup', content=content,
              size=(400,200), size_hint=(None, None))

        popup.open()

    def backup_copy(self, path, status):
        if os.path.isdir(path.text):
            copy(self.user_sheet_path, path.text)
            copy(self.aloc_sheet_path, path.text)
            status.text = 'Arquivos copiados com sucesso!'
            status.color = (0,1,0,1)
        else:
            status.text = 'Caminho especificado não existe.'
            status.color = (1,0,0,1)
        #copy(self.user_sheet_path, 'bk')
        #copy(self.aloc_sheet_path, 'bk')
        
class UserDialog(BoxLayout):
    alocar = ObjectProperty()
    salvar_user = ObjectProperty()
    cancel = ObjectProperty()
    user_nome = ObjectProperty()
    user_doc = ObjectProperty()
    user_placas = ObjectProperty()
    linha_planilha = ObjectProperty()
    data_cadastro = ObjectProperty()
    texto_regex = re.compile('[A-Z a-z]*')
    placas_regex = re.compile('(([A-Z]{3}-{1}\d{4}(;|$))|((?=(?:\d*[A-Z]){4})(?=(?:[A-Z]*\d){3})[\w\d]{7}(;|$))){1,}')

    def iter_rows(self, ws):
        row_values = []
        for line, row in enumerate(ws.iter_rows(), 1):
            if str(line) != self.linha_planilha.split(': ')[1]:
                yield row[1].value

    def valid_texto(self, texto):
        result = self.texto_regex.fullmatch(texto)
        return result

    def valid_doc(self, texto):
        MyApp.carregar_sheet_user()

        for doc in self.iter_rows(MyApp.user_sheet):
            if str(doc) == texto:
                MyApp.fechar_sheet_user()
                return False
        MyApp.fechar_sheet_user()
        return True

    def valid_placas(self, placas):
        placas = placas.upper()
        result = self.placas_regex.fullmatch(placas)
        return result

    def check_valores(self, nome, doc, placas):
        self.ids.user_nome.text = self.ids.user_nome.text.title()
        self.ids.user_placas.text = self.ids.user_placas.text.upper()

        if len(doc) > 14: self.ids.user_doc.text = doc[0:14]
        if len(nome) > 75: self.ids.user_nome.text = nome[0:75]
        if len(placas) > 50: self.ids.user_placas.text = placas[0:50]
            
        self.ids.salvar_user.disabled = True

        self.label_ver_dados.color = (1,0,0,1)
        self.ids.user_nome.foreground_color = (0, 0, 0, 1)
        self.ids.user_doc.foreground_color = (0, 0, 0, 1)
        self.ids.user_placas.foreground_color = (0, 0, 0, 1)
        if nome and not self.valid_texto(nome):
            self.label_ver_dados.text = 'Nome deve conter apenas letras.' 
            self.ids.user_nome.foreground_color = (1, 0, 0, 1)
            return
        elif doc and not self.valid_doc(doc):
            self.label_ver_dados.text = 'Documento já registrado.\nFaça busca antes de registrar novo usuário.' 
            self.ids.user_doc.foreground_color = (1, 0, 0, 1)
            return
        elif placas and not self.valid_placas(placas):
            self.label_ver_dados.text = 'Padrão: AAA-1234. Se houver mais de uma placa separe-as por ponto e vírgula ;\nPadrão Mercosul: 4 letras e 3 números em qualquer ordem.' 
            self.ids.user_placas.foreground_color = (1, 0, 0, 1)
            return
        elif not nome or not doc or nome=='' or doc=='':
            self.label_ver_dados.text = 'Preencher Nome e Doc.'
            return

        self.ids.salvar_user.disabled = False
        self.label_ver_dados.color = (0,1,0,1)
        self.label_ver_dados.text = 'Valores válidos!'

    def salvar_user(self, result, popup, content):
        global MyApp
        try:
            MyApp.carregar_sheet_user()
            MyApp.user_sheet.cell(row=int(result[4]), column=1).value = content.ids.user_nome.text
            MyApp.user_sheet.cell(row=int(result[4]), column=2).value = content.ids.user_doc.text
            MyApp.user_sheet.cell(row=int(result[4]), column=3).value = content.ids.user_placas.text
            MyApp.all_sheets_user.save(MyApp.user_sheet_path)
            MyApp.load_usertable_results(MyApp.buscar_user.text)

            popup.dismiss()
            MyApp.fechar_sheet_user()
        except:
            traceback.print_exc(file=sys.stdout)
            msg = Label(text='Se estiver com a planilha aberta, feche e tente novamente.')
            popup = Popup(title='Erro',
                          content=msg,
                          size_hint=(0.3, 0.25))
            popup.open()

class NovoUser(BoxLayout):
    add_user = ObjectProperty()
    cancel = ObjectProperty()
    user_nome = ObjectProperty()
    user_doc = ObjectProperty()
    user_placas = ObjectProperty()
    label_ver_dados = ObjectProperty()
    data_cadastro = ObjectProperty()
    texto_regex = re.compile('[A-Z a-z]*')
    placas_regex = re.compile('(([A-Z]{3}-{1}\d{4}(;|$))|((?=(?:\d*[A-Z]){4})(?=(?:[A-Z]*\d){3})[\w\d]{7}(;|$))){1,}')

    def iter_rows(self, ws):
        row_values = []
        for row in ws.iter_rows():
            yield row[1].value

    def valid_texto(self, texto):
        result = self.texto_regex.fullmatch(texto)
        return result

    def valid_doc(self, texto):
        MyApp.carregar_sheet_user()

        for doc in self.iter_rows(MyApp.user_sheet):
            if str(doc) == texto:
                MyApp.fechar_sheet_user()
                return False
        MyApp.fechar_sheet_user()
        return True

    def valid_placas(self, placas):
        placas = placas.upper()
        result = self.placas_regex.fullmatch(placas)
        return result

    def check_valores(self, nome, doc, placas):
        self.user_nome.text = self.user_nome.text.title()
        self.user_placas.text = self.user_placas.text.upper()

        if len(doc) > 14: self.user_doc.text = doc[0:14]
        if len(nome) > 75: self.user_nome.text = nome[0:75]
        if len(placas) > 50: self.user_placas.text = placas[0:50]
            
        self.add_user.disabled = True

        self.label_ver_dados.color = (1,0,0,1)
        self.user_nome.foreground_color = (0, 0, 0, 1)
        self.user_doc.foreground_color = (0, 0, 0, 1)
        self.user_placas.foreground_color = (0, 0, 0, 1)
        if nome and not self.valid_texto(nome):
            self.label_ver_dados.text = 'Nome deve conter apenas letras.' 
            self.user_nome.foreground_color = (1, 0, 0, 1)
            return
        elif doc and not self.valid_doc(doc):
            self.label_ver_dados.text = 'Documento já registrado.\nFaça busca antes de registrar novo usuário.' 
            self.user_doc.foreground_color = (1, 0, 0, 1)
            return
        elif placas and not self.valid_placas(placas):
            self.label_ver_dados.text = 'Padrão: AAA-1234. Se houver mais de uma placa separe-as por ponto e vírgula ;\nPadrão Mercosul: 4 letras e 3 números em qualquer ordem.' 
            self.user_placas.foreground_color = (1, 0, 0, 1)
            return
        elif not nome or not doc or nome=='' or doc=='':
            self.label_ver_dados.text = 'Preencher Nome e Doc.'
            return

        self.add_user.disabled = False
        self.label_ver_dados.color = (0,1,0,1)
        self.label_ver_dados.text = 'Valores válidos!'

    def adicionar(self, popup):
        global MyApp
        try:
            MyApp.carregar_sheet_user()
            row_n=MyApp.sheet_len(MyApp.user_sheet)+1

            MyApp.user_sheet.cell(row=row_n, column=1).value = self.user_nome.text
            MyApp.user_sheet.cell(row=row_n, column=2).value = self.user_doc.text
            MyApp.user_sheet.cell(row=row_n, column=3).value = self.user_placas.text
            MyApp.user_sheet.cell(row=row_n, column=4).value = self.data_cadastro.split(': ')[1]
            MyApp.all_sheets_user.save(MyApp.user_sheet_path)

            MyApp.buscar_user.text = self.user_nome.text
            MyApp.load_usertable_results(MyApp.buscar_user.text)
            MyApp.fechar_sheet_user()
        except:
            traceback.print_exc(file=sys.stdout)
            msg = Label(text='Se estiver com a planilha aberta, feche e tente novamente.')
            popup_error = Popup(title='Erro',
                                content=msg,
                                size_hint=(0.5, 0.25))
            popup_error.open()
        popup.dismiss()


class AlocDialog(BoxLayout):
    salvar_aloc = ObjectProperty()
    cancel = ObjectProperty()
    linha_planilha = ObjectProperty()
    user_nome = ObjectProperty()
    user_doc = ObjectProperty()
    user_placas = ObjectProperty()
    data_entrada = ObjectProperty()
    data_saida = ObjectProperty()
    liberado = ObjectProperty()
    preco = ObjectProperty()
    digitos = ''

    def __init__(self, **kwargs):
        super(AlocDialog, self).__init__(**kwargs)
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)

    def _keyboard_closed(self):
        self._keyboard = None

    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        if len(self.digitos) >= 8:
            if keycode[1] == 'backspace':
                self.digitos = self.digitos[:-1]
            return
        if len(self.digitos) > 0 and self.digitos[0] == '.':
            self.digitos = ''
            return
        if '.' in self.digitos:
            if keycode[1] == 'backspace':
                if len(self.digitos.split('.')[1]) >= 2:
                    self.digitos = self.digitos[:-1]
                elif len(self.digitos.split('.')[1]) == 1:
                    self.digitos = self.digitos[:-2]
                else:
                    self.digitos = self.digitos[:-1]
            elif len(self.digitos.split('.')[1]) < 2 and keycode[1][6:] in ['0','1','2','3','4','5','6','7','8','9']:
                self.digitos = self.digitos + keycode[1][-1:]
            else:
                return
        elif keycode[1] == 'backspace':
            self.digitos = self.digitos[:-1]
        elif 'decimal' in keycode[1] or keycode[1] in ['.', ',']:
            self.digitos = self.digitos + '.'
        elif 'numpad' in keycode[1] and keycode[1][6:] in ['0','1','2','3','4','5','6','7','8','9']:
            self.digitos = self.digitos + keycode[1][-1:]
        else:
            return

        return True

    def set_cursor(self, instance):
        self.preco.do_cursor_movement('cursor_right')

    def moeda_format(self, instance):
        if len(self.digitos) > 0 and self.digitos[0] == '.':
            self.digitos = ''
            return
        if self.digitos:
            locale.setlocale(locale.LC_ALL, 'en_CA.UTF-8')
            preco = re.sub("[^0-9\.]", "", self.digitos)

            ver = locale.currency(float(preco), grouping=True)
            self.preco.text = ver
            Clock.schedule_once(self.set_cursor, 0)
        else:
            self.preco.text = '$0.00'

    def liberar_aloc(self, popup):
        linha = self.linha_planilha.split(': ')[1]
        try:
            MyApp.carregar_sheet_aloc()

            if self.ids.liberado.text == 'SIM':
                MyApp.aloc_sheet.cell(row=int(linha), column=5).value = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            else:
                MyApp.aloc_sheet.cell(row=int(linha), column=5).value = 'NÃO'
            MyApp.aloc_sheet.cell(row=int(linha), column=6).value = self.ids.preco.text
            MyApp.aloc_sheet.cell(row=int(linha), column=7).value = self.ids.liberado.text

            MyApp.all_sheets_aloc.save(MyApp.aloc_sheet_path)
            MyApp.load_aloctable_results(MyApp.buscar_aloc.text)
            MyApp.fechar_sheet_aloc()
        except:
            traceback.print_exc(file=sys.stdout)
            msg = Label(text='Se estiver com a planilha aberta, feche e tente novamente.')
            popup = Popup(title='Erro',
                          content=msg,
                          size_hint=(0.3, 0.25))
            popup.open()
        popup.dismiss()

class NovoAloc(BoxLayout):
    salvar_aloc = ObjectProperty()
    cancel = ObjectProperty()
    user_nome = ObjectProperty()
    user_doc = ObjectProperty()
    user_placas = ObjectProperty()
    data_entrada = ObjectProperty()
    preco = ObjectProperty()
    digitos = ''

    def __init__(self, **kwargs):
        super(NovoAloc, self).__init__(**kwargs)
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)

    def _keyboard_closed(self):
        self._keyboard = None

    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        if len(self.digitos) >= 8:
            if keycode[1] == 'backspace':
                self.digitos = self.digitos[:-1]
            return
        if len(self.digitos) > 0 and self.digitos[0] == '.':
            self.digitos = ''
            return
        if '.' in self.digitos:
            if keycode[1] == 'backspace':
                if len(self.digitos.split('.')[1]) >= 2:
                    self.digitos = self.digitos[:-1]
                elif len(self.digitos.split('.')[1]) == 1:
                    self.digitos = self.digitos[:-2]
                else:
                    self.digitos = self.digitos[:-1]
            elif len(self.digitos.split('.')[1]) < 2 and keycode[1][6:] in ['0','1','2','3','4','5','6','7','8','9']:
                self.digitos = self.digitos + keycode[1][-1:]
            else:
                return
        elif keycode[1] == 'backspace':
            self.digitos = self.digitos[:-1]
        elif 'decimal' in keycode[1] or keycode[1] in ['.', ',']:
            self.digitos = self.digitos + '.'
        elif 'numpad' in keycode[1] and keycode[1][6:] in ['0','1','2','3','4','5','6','7','8','9']:
            self.digitos = self.digitos + keycode[1][-1:]
        else:
            return

        return True

    def set_cursor(self, instance):
        self.preco.do_cursor_movement('cursor_right')

    def moeda_format(self, instance):
        if len(self.digitos) > 0 and self.digitos[0] == '.':
            self.digitos = ''
            return
        if self.digitos:
            locale.setlocale(locale.LC_ALL, 'en_CA.UTF-8')
            preco = re.sub("[^0-9\.]", "", self.digitos)

            ver = locale.currency(float(preco), grouping=True)
            self.preco.text = ver
            Clock.schedule_once(self.set_cursor, 0)
        else:
            self.preco.text = '$0.00'

    def adicionar_aloc(self, popup):
        #linha = self.linha_planilha.split(': ')[1]
        try:
            MyApp.carregar_sheet_aloc()
            row_n=MyApp.sheet_len(MyApp.aloc_sheet)+1
            MyApp.aloc_sheet.cell(row=row_n, column=1).value = self.user_nome
            MyApp.aloc_sheet.cell(row=row_n, column=2).value = self.user_doc
            MyApp.aloc_sheet.cell(row=row_n, column=3).value = self.ids.user_placas.text
            MyApp.aloc_sheet.cell(row=row_n, column=4).value = self.data_entrada.split(':\n')[1]
            #MyApp.aloc_sheet.cell(row=row_n, column=5).value = ''
            MyApp.aloc_sheet.cell(row=row_n, column=6).value = self.preco.text
            MyApp.aloc_sheet.cell(row=row_n, column=7).value = 'NÃO'

            MyApp.all_sheets_aloc.save(MyApp.aloc_sheet_path)
            MyApp.load_aloctable_results(self.user_doc)
            MyApp.fechar_sheet_aloc()
        except:
            traceback.print_exc(file=sys.stdout)
            msg = Label(text='Se estiver com a planilha aberta, feche e tente novamente.')
            popup = Popup(title='Erro',
                          content=msg,
                          size_hint=(0.3, 0.25))
            popup.open()
        popup.dismiss()

class Parking(App):

    def build(self):
        global MyApp
        self.title = 'Estacionamento'
        Builder.load_string(open("UI.kv", encoding="utf-8").read(), rulesonly=True)
        MyApp = UI()
        Window.bind(on_request_close=self.on_request_close)
        return MyApp

    def on_request_close(self, *args):
        self.backup_planilhas()
        self.stop()
        return True

    def backup_planilhas(self):
        global MyApp
        copy(MyApp.user_sheet_path, 'bk')
        copy(MyApp.aloc_sheet_path, 'bk')
        

Factory.register('UserDialog', cls=UserDialog)
Factory.register('NovoUser', cls=NovoUser)
Factory.register('AlocDialog', cls=AlocDialog)

if __name__ == '__main__':
    Parking().run()