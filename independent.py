import unicodedata
import webbrowser
import openpyxl
from datetime import date, datetime
from kivy.clock import Clock
from kivy.config import Config

def screensize_conf(config):
    if not config['Default']['width'] or not config['Default']['height']:
        import ctypes
        user32 = ctypes.windll.user32
        screensize = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
        config['Default']['width'] = str(user32.GetSystemMetrics(0)-200)
        config['Default']['height'] = str(user32.GetSystemMetrics(1)-150)
        with open('config.ini', 'w') as configfile: #save
            config.write(configfile)

def screensize_set(config):
    Config.set('graphics', 'width', config['Default']['width'])
    Config.set('graphics', 'height', config['Default']['height'])
    Config.set('graphics','borderless', 0)
    Config.set('graphics','resizable', 1)
    #Config.set('input', 'mouse', 'mouse,multitouch_on_demand')
    Config.set('kivy', 'exit_on_escape', '1')
    Config.write()

def remove_accents(input_str):
    """
    Remove acentuacao para realizar buscas
    """
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    only_ascii = nfkd_form.encode('ASCII', 'ignore')
    return only_ascii

def sheet_len(ws):
    """
    Retorna a quantidade de linhas usadas na planilha
    """
    res = 0
    for row in ws.iter_rows():
        x = [cell.value for cell in row]
        if not all(s == None for s in x):
            res += 1
    return res

def iter_rows(ws):
    """
    Retorna uma lista de listas com o conteudo de cada linha
    ocupada na planilha e reverte para mostrar os ultimos
    cadastros primeiro
    """
    row_values = []
    for row in ws.iter_rows():
        x = [cell.value for cell in row]
        if not all(s == None for s in x):
            row_values.append(x)
    return reversed(row_values)

def convert_to_string(vlist):
    """
    Converte para numero para string
    """
    new_list = []
    for v in vlist:
        if isinstance(v, float):
            new_list.append(str(int(v)))
        elif not v:
            new_list.append('')
        else:
            new_list.append(str(v))
    return new_list

def set_cursor_home(cursor, instance):
    """
    Coloca o cursor na primeira posicao no campo de texto.
    Necessario para tornar o inicio do texto visivel.
    """
    cursor.do_cursor_movement('cursor_home')

def set_cursor_right(cursor, instance):
    """
    Coloca o cursor na ultima posicao no campo de texto.
    Necessario para digitar campo de dinheiro.
    """
    cursor.do_cursor_movement('cursor_right')

def _refocus_text_input(field, instance):
    """
    Mantem o foco no campo de texto depois de uma busca
    """
    field.focus=True

def atualiza_horario(content, tipo, instance):
    """
    Usado para atualizar o horario no campo de texto em tempo real
    """
    if tipo == 'add_aloc':
        content.data_entrada=str(12*' ')+'Entrada:\n'+datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    elif tipo == 'add_user':
        content.data_cadastro=str('Data de cadastro: '+datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

def unclock(event, popup):
    """
    Usado para desativar um clock schedule
    """
    Clock.unschedule(event)
    popup.dismiss()

def open_link(instance, link):
    webbrowser.open(link)

def carregar_sheet_user(user_sheet_path):
    all_sheets_user = openpyxl.load_workbook(user_sheet_path)
    return all_sheets_user, all_sheets_user[all_sheets_user.sheetnames[0]]

def carregar_sheet_aloc(aloc_sheet_path):
    all_sheets_aloc = openpyxl.load_workbook(aloc_sheet_path)
    return all_sheets_aloc, all_sheets_aloc[all_sheets_aloc.sheetnames[0]]

def fechar_sheet(all_sheets):
    all_sheets.close()